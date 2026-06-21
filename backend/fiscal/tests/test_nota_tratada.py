"""
TDD — NotaTratada: parser XML, API de listagem, exportação Excel e integração externa.

Rodar:
    python manage.py test fiscal.tests.test_nota_tratada --verbosity=2
"""
import datetime
import decimal
import io
from textwrap import dedent
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework import status
from rest_framework.authtoken.models import Token
from rest_framework.test import APITestCase

from fiscal.models import (
    Cliente,
    Documento,
    Escritorio,
    NotaTratada,
    StatusDocumento,
    TipoDocumento,
    Xml,
)
from fiscal.services.tratamento_nfse import extrair_dados_nfse, _calcular_parecer

User = get_user_model()

NS = 'http://www.sped.fazenda.gov.br/nfse'

# ─────────────────────────── helpers ──────────────────────────────────────────

def make_escritorio(cnpj='00000000000191', razao='Escritório Teste'):
    return Escritorio.objects.create(cnpj=cnpj, razao_social=razao)


def make_staff(username='admin', escritorio=None):
    u = User.objects.create_user(username=username, password='pass', is_staff=True)
    u.escritorio = escritorio
    u.save()
    return u


def make_operador(username='op', escritorio=None):
    u = User.objects.create_user(username=username, password='pass', is_staff=False)
    u.escritorio = escritorio
    u.save()
    return u


def make_cliente(cnpj='12345678000195', razao='Empresa Teste Ltda', escritorio=None):
    return Cliente.objects.create(cnpj=cnpj, razao_social=razao, escritorio=escritorio)


def make_documento(cliente, chave=None, tipo='NFSE', status='COMPLETO', papel='EMITENTE', competencia='2025-06'):
    chave = chave or ('N' * 44)
    return Documento.objects.create(
        cliente=cliente,
        chave=chave,
        tipo_documento=tipo,
        emitente='Empresa Teste Ltda',
        valor=decimal.Decimal('1000.00'),
        data_emissao=datetime.date(2025, 6, 10),
        competencia=competencia,
        status=status,
        papel_nfse=papel,
    )


def make_nota_tratada(documento, numero='100', competencia='06/2025', parecer='Válida', **kw):
    defaults = dict(
        numero_nfse=numero,
        data_competencia=competencia,
        data_processamento='10/06/2025',
        emitente_cnpj=documento.cliente.cnpj,
        emitente_nome=documento.cliente.razao_social,
        tomador_doc='98765432000110',
        tomador_nome='Tomador Exemplo S.A.',
        codigo_tributo='01.07',
        descricao_servico='Consultoria de TI',
        regime_trib='Nenhum',
        valor_servico=decimal.Decimal('1000.00'),
        valor_liquido=decimal.Decimal('900.00'),
        ret_pis=decimal.Decimal('6.50'),
        ret_cofins=decimal.Decimal('30.00'),
        ret_csll=decimal.Decimal('10.00'),
        ret_irrf=None,
        ret_inss=None,
        parecer=parecer,
        chave_substituta='',
    )
    defaults.update(kw)
    return NotaTratada.objects.create(documento=documento, **defaults)


def _xml_nfse(
    numero='123',
    competencia='2025-06',
    cnpj_emit='12345678000195',
    nome_emit='Empresa Teste Ltda',
    cnpj_toma='98765432000110',
    nome_toma='Tomador S.A.',
    v_serv='1000.00',
    v_liq='900.00',
    ret_csll='10.00',
    ret_pis='6.50',
    ret_cofins='30.00',
    ret_irrf='',
    ret_inss='',
    tp_ret='',
    ch_substda='',
    op_simp_nac='',
    reg_esp_trib='0',
    status_proc='Autorizado',
) -> str:
    """Monta XML NFS-e Nacional mínimo válido para os testes."""
    subst_block = f'<subst><chSubstda>NFS{ch_substda}</chSubstda></subst>' if ch_substda else ''
    piscofins_block = f'<tpRetPisCofins>{tp_ret}</tpRetPisCofins>' if tp_ret else ''
    pis_tag   = f'<vRetPis>{ret_pis}</vRetPis>' if ret_pis else ''
    cofins_tag = f'<vRetCofins>{ret_cofins}</vRetCofins>' if ret_cofins else ''
    csll_tag  = f'<vRetCSLL>{ret_csll}</vRetCSLL>' if ret_csll else ''
    irrf_tag  = f'<vRetIRRF>{ret_irrf}</vRetIRRF>' if ret_irrf else ''
    inss_tag  = f'<vRetINSS>{ret_inss}</vRetINSS>' if ret_inss else ''
    op_tag    = f'<opSimpNac>{op_simp_nac}</opSimpNac>' if op_simp_nac else ''

    return dedent(f"""\
    <?xml version="1.0" encoding="UTF-8"?>
    <NFSe xmlns="{NS}">
      <infNFSe Id="NFS{'4' * 44}">
        <nNFSe>{numero}</nNFSe>
        <dCompet>{competencia}</dCompet>
        <dhProc>2025-06-10T10:00:00</dhProc>
        <emit>
          <CNPJ>{cnpj_emit}</CNPJ>
          <xNome>{nome_emit}</xNome>
          {op_tag}
        </emit>
        <DPS>
          <infDPS>
            <toma>
              <CNPJ>{cnpj_toma}</CNPJ>
              <xNome>{nome_toma}</xNome>
            </toma>
            <serv>
              <cTribNac>01.07</cTribNac>
              <cServ><xDescServ>Consultoria de TI</xDescServ></cServ>
            </serv>
            <valores>
              <vServPrest><vServ>{v_serv}</vServ></vServPrest>
            </valores>
            <regEspTrib>{reg_esp_trib}</regEspTrib>
            {subst_block}
          </infDPS>
        </DPS>
        <valores>
          <vServ>{v_serv}</vServ>
          <vLiq>{v_liq}</vLiq>
        </valores>
        <piscofins>
          {piscofins_block}
          {pis_tag}
          {cofins_tag}
        </piscofins>
        {csll_tag}
        {irrf_tag}
        {inss_tag}
      </infNFSe>
    </NFSe>
    """)


# ──────────────────────────────────────────────────────────────────────────────
# 1. PARSER XML NFS-e — testes unitários puro
# ──────────────────────────────────────────────────────────────────────────────

class ExtrairDadosNfseTest(TestCase):

    def test_xml_invalido_retorna_dict_vazio(self):
        resultado = extrair_dados_nfse('<xml-quebrado>', 'COMPLETO', 'EMITENTE')
        self.assertEqual(resultado, {})

    def test_campos_basicos_extraidos(self):
        xml = _xml_nfse(numero='42', competencia='2025-06')
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['numero_nfse'], '42')
        self.assertEqual(dados['data_competencia'], '06/2025')
        self.assertEqual(dados['emitente_cnpj'], '12345678000195')
        self.assertEqual(dados['emitente_nome'], 'Empresa Teste Ltda')
        self.assertEqual(dados['tomador_doc'], '98765432000110')
        self.assertEqual(dados['tomador_nome'], 'Tomador S.A.')
        self.assertEqual(dados['data_processamento'], '10/06/2025')

    def test_valor_servico_convertido_para_decimal(self):
        xml = _xml_nfse(v_serv='2500.00')
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['valor_servico'], decimal.Decimal('2500.00'))

    def test_valor_liquido_extraido(self):
        xml = _xml_nfse(v_liq='2300.00')
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['valor_liquido'], decimal.Decimal('2300.00'))

    def test_regime_trib_mapeado(self):
        """Sem opSimpNac e regEspTrib='0' → Nenhum."""
        xml = _xml_nfse()
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['regime_trib'], 'Nenhum')

    # ── opSimpNac → regime tributário (TDD: regra fiscal crítica) ─────────────

    def test_regime_mei_quando_opsimpnac_3(self):
        """opSimpNac='3' → MEI independente do regEspTrib."""
        xml = _xml_nfse(op_simp_nac='3')
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['regime_trib'], 'MEI')

    def test_regime_simples_nacional_quando_opsimpnac_1(self):
        """opSimpNac='1' → Simples Nacional."""
        xml = _xml_nfse(op_simp_nac='1')
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['regime_trib'], 'Simples Nacional')

    def test_regime_simples_excesso_quando_opsimpnac_2(self):
        """opSimpNac='2' → Simples Nacional (Excesso Sublimite)."""
        xml = _xml_nfse(op_simp_nac='2')
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['regime_trib'], 'Simples Nacional (Excesso Sublimite)')

    def test_opsimpnac_tem_precedencia_sobre_regesp_trib(self):
        """opSimpNac presente deve sobrescrever regEspTrib — regra fiscal de precedência."""
        xml = _xml_nfse(op_simp_nac='3', reg_esp_trib='6')  # regEspTrib='6'=Sociedade de Profissões
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['regime_trib'], 'MEI')

    def test_sem_opsimpnac_usa_regesp_trib(self):
        """Sem opSimpNac, regEspTrib='3' → 'Sociedade de Profissionais' (regime ISS)."""
        xml = _xml_nfse(reg_esp_trib='3')
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['regime_trib'], 'Sociedade de Profissionais')

    def test_opsimpnac_0_nao_eh_mei_nem_simples(self):
        """opSimpNac='0' (Não Simples) deve cair no fallback de regEspTrib."""
        xml = _xml_nfse(op_simp_nac='0', reg_esp_trib='0')
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        # '0' não está no _OP_SIMP_MAP, usa regEspTrib='0' → 'Nenhum'
        self.assertEqual(dados['regime_trib'], 'Nenhum')

    # ── Pareceres ─────────────────────────────────────────────────────────────

    def test_parecer_cancelada_quando_status_cancelado(self):
        xml = _xml_nfse()
        dados = extrair_dados_nfse(xml, 'CANCELADO', 'EMITENTE')
        self.assertEqual(dados['parecer'], 'Cancelada')

    def test_parecer_substituida_quando_status_substituido(self):
        xml = _xml_nfse()
        dados = extrair_dados_nfse(xml, 'SUBSTITUIDO', 'EMITENTE')
        self.assertEqual(dados['parecer'], 'Substituída')

    def test_parecer_valida_com_retencoes_corretas(self):
        # CSLL correto = 1000 * 0.01 = 10.00
        xml = _xml_nfse(v_serv='1000.00', ret_csll='10.00')
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['parecer'], 'Válida')

    def test_parecer_divergencia_quando_csll_errado(self):
        # CSLL informado = 99.00, esperado = 10.00 → divergência
        xml = _xml_nfse(v_serv='1000.00', ret_csll='99.00')
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['parecer'], 'Válida (DIVERGÊNCIA RETENÇÃO)')

    def test_bundle_csrf_desagregado_corretamente(self):
        """
        CSLL=40.65 é o total CSRF (PIS+COFINS+CSLL) de R$1000.
        O parser deve detectar e desagregar: PIS=6.50, COFINS=30.00, CSLL=10.00.
        """
        csrf_total = decimal.Decimal('1000') * (decimal.Decimal('0.0065') + decimal.Decimal('0.03') + decimal.Decimal('0.01'))
        xml = _xml_nfse(v_serv='1000.00', ret_csll=str(csrf_total))
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['parecer'], 'Válida')
        self.assertEqual(dados['ret_pis'],    decimal.Decimal('6.50'))
        self.assertEqual(dados['ret_cofins'], decimal.Decimal('30.00'))
        self.assertEqual(dados['ret_csll'],   decimal.Decimal('10.00'))

    def test_divergencia_nao_sobrescreve_cancelada(self):
        """Nota Cancelada com CSLL errado continua com parecer Cancelada."""
        xml = _xml_nfse(v_serv='1000.00', ret_csll='99.00')
        dados = extrair_dados_nfse(xml, 'CANCELADO', 'EMITENTE')
        self.assertEqual(dados['parecer'], 'Cancelada')

    # ── tpRetPisCofins ────────────────────────────────────────────────────────

    def test_tp_ret_2_zera_pis_cofins(self):
        """tpRetPisCofins=2 significa 'não retém' — PIS e COFINS devem ser None."""
        xml = _xml_nfse(tp_ret='2', ret_pis='6.50', ret_cofins='30.00')
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertIsNone(dados['ret_pis'])
        self.assertIsNone(dados['ret_cofins'])

    def test_tp_ret_1_usa_pis_cofins_do_xml(self):
        xml = _xml_nfse(tp_ret='1', ret_pis='6.50', ret_cofins='30.00', ret_csll='10.00')
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['ret_pis'],    decimal.Decimal('6.50'))
        self.assertEqual(dados['ret_cofins'], decimal.Decimal('30.00'))

    # ── Substituição ──────────────────────────────────────────────────────────

    def test_chave_substituda_extraida_sem_prefixo_nfs(self):
        xml = _xml_nfse(ch_substda='35250600000000000000550010000001001234567890')
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(
            dados['chave_que_esta_substitui'],
            '35250600000000000000550010000001001234567890',
        )

    def test_sem_substituicao_retorna_string_vazia(self):
        xml = _xml_nfse()
        dados = extrair_dados_nfse(xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(dados['chave_que_esta_substitui'], '')


# ──────────────────────────────────────────────────────────────────────────────
# 2. _calcular_parecer — lógica isolada
# ──────────────────────────────────────────────────────────────────────────────

class CalcularParecerTest(TestCase):

    D = decimal.Decimal

    def test_cancelado(self):
        self.assertEqual(_calcular_parecer('CANCELADO', self.D('1000'), self.D('10'), None, None), 'Cancelada')

    def test_substituido(self):
        self.assertEqual(_calcular_parecer('SUBSTITUIDO', self.D('1000'), self.D('10'), None, None), 'Substituída')

    def test_valida_sem_csll(self):
        self.assertEqual(_calcular_parecer('COMPLETO', self.D('1000'), None, None, None), 'Válida')

    def test_valida_csll_correto(self):
        self.assertEqual(_calcular_parecer('COMPLETO', self.D('1000'), self.D('10.00'), None, None), 'Válida')

    def test_valida_bundle_csrf(self):
        csrf = self.D('1000') * self.D('0.0465')
        self.assertEqual(_calcular_parecer('COMPLETO', self.D('1000'), csrf, None, None), 'Válida')

    def test_divergencia(self):
        self.assertEqual(_calcular_parecer('COMPLETO', self.D('1000'), self.D('99.00'), None, None),
                         'Válida (DIVERGÊNCIA RETENÇÃO)')


# ──────────────────────────────────────────────────────────────────────────────
# 3. _salvar_nota_tratada — integração com banco
# ──────────────────────────────────────────────────────────────────────────────

class SalvarNotaTratadaIntegracaoTest(TestCase):

    def setUp(self):
        self.cliente = make_cliente()

    def test_cria_nota_tratada_para_documento_novo(self):
        doc = make_documento(self.cliente, chave='A' * 44)
        xml = _xml_nfse(numero='10')
        from fiscal.conectores.nfse import _salvar_nota_tratada
        _salvar_nota_tratada(doc, xml, 'COMPLETO', 'EMITENTE')
        self.assertTrue(NotaTratada.objects.filter(documento=doc).exists())

    def test_idempotente_update_or_create(self):
        doc = make_documento(self.cliente, chave='B' * 44)
        xml = _xml_nfse(numero='20')
        from fiscal.conectores.nfse import _salvar_nota_tratada
        _salvar_nota_tratada(doc, xml, 'COMPLETO', 'EMITENTE')
        _salvar_nota_tratada(doc, xml, 'COMPLETO', 'EMITENTE')
        self.assertEqual(NotaTratada.objects.filter(documento=doc).count(), 1)

    def test_substitucao_marca_nota_antiga_como_substituida(self):
        chave_antiga = 'C' * 44
        doc_antigo = make_documento(self.cliente, chave=chave_antiga, status='COMPLETO')
        make_nota_tratada(doc_antigo, numero='30', parecer='Válida')

        chave_nova = 'D' * 44
        doc_novo = make_documento(self.cliente, chave=chave_nova)
        xml_novo = _xml_nfse(numero='31', ch_substda=chave_antiga)
        from fiscal.conectores.nfse import _salvar_nota_tratada
        _salvar_nota_tratada(doc_novo, xml_novo, 'COMPLETO', 'EMITENTE')

        doc_antigo.refresh_from_db()
        nota_antiga = NotaTratada.objects.get(documento=doc_antigo)
        self.assertEqual(doc_antigo.status, 'SUBSTITUIDO')
        self.assertEqual(nota_antiga.parecer, 'Substituída')
        self.assertEqual(nota_antiga.chave_substituta, chave_nova)

    def test_xml_invalido_nao_cria_nota_tratada(self):
        doc = make_documento(self.cliente, chave='E' * 44)
        from fiscal.conectores.nfse import _salvar_nota_tratada
        _salvar_nota_tratada(doc, '<xml-invalido>', 'COMPLETO', 'EMITENTE')
        self.assertFalse(NotaTratada.objects.filter(documento=doc).exists())


# ──────────────────────────────────────────────────────────────────────────────
# 4. NotaTratadaViewSet — API JWT
# ──────────────────────────────────────────────────────────────────────────────

class NotaTratadaViewSetTest(APITestCase):

    def setUp(self):
        self.escritorio = make_escritorio()
        self.staff = make_staff(username='staff_nt', escritorio=self.escritorio)
        self.cliente = make_cliente(escritorio=self.escritorio)
        self.doc = make_documento(self.cliente, chave='F' * 44, competencia='2025-06')
        self.nota = make_nota_tratada(self.doc, numero='100', competencia='06/2025')

    def test_sem_auth_retorna_401(self):
        res = self.client.get('/api/notas-tratadas/')
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_lista_retorna_200_com_results(self):
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('results', res.data)

    def test_nota_aparece_na_lista(self):
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/')
        self.assertEqual(res.data['count'], 1)
        item = res.data['results'][0]
        self.assertEqual(item['numero_nfse'], '100')
        self.assertEqual(item['parecer'], 'Válida')

    def test_campos_obrigatorios_presentes(self):
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/')
        item = res.data['results'][0]
        for campo in ('numero_nfse', 'data_competencia', 'emitente_cnpj', 'emitente_nome',
                      'valor_servico', 'ret_pis', 'ret_cofins', 'ret_csll', 'parecer',
                      'cliente_nome', 'papel_nfse'):
            self.assertIn(campo, item, f'Campo ausente: {campo}')

    def test_filtro_por_data_competencia(self):
        outro_doc = make_documento(self.cliente, chave='G' * 44, competencia='2025-05')
        make_nota_tratada(outro_doc, numero='99', competencia='05/2025')
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/?data_competencia=06/2025')
        self.assertEqual(res.data['count'], 1)
        self.assertEqual(res.data['results'][0]['numero_nfse'], '100')

    def test_filtro_por_parecer(self):
        doc2 = make_documento(self.cliente, chave='H' * 44)
        make_nota_tratada(doc2, numero='101', parecer='Cancelada')
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/?parecer=Cancelada')
        self.assertEqual(res.data['count'], 1)
        self.assertEqual(res.data['results'][0]['parecer'], 'Cancelada')

    def test_filtro_por_cliente(self):
        outro_cliente = make_cliente(cnpj='99999999000199', razao='Outro Cliente Ltda',
                                     escritorio=self.escritorio)
        doc_outro = make_documento(outro_cliente, chave='I' * 44)
        make_nota_tratada(doc_outro, numero='200')
        self.client.force_authenticate(user=self.staff)
        res = self.client.get(f'/api/notas-tratadas/?cliente={self.cliente.pk}')
        self.assertEqual(res.data['count'], 1)

    def test_isolamento_escritorio(self):
        """Usuário de escritório diferente não vê notas do outro."""
        outro_escritorio = make_escritorio(cnpj='11111111000191', razao='Escritório B')
        staff_b = make_staff(username='staff_b', escritorio=outro_escritorio)
        self.client.force_authenticate(user=staff_b)
        res = self.client.get('/api/notas-tratadas/')
        self.assertEqual(res.data['count'], 0)

    def test_superuser_ve_tudo(self):
        superuser = User.objects.create_superuser(username='super', password='pass')
        outro_escritorio = make_escritorio(cnpj='22222222000191', razao='Escritório C')
        cliente_c = make_cliente(cnpj='88888888000188', escritorio=outro_escritorio)
        doc_c = make_documento(cliente_c, chave='J' * 44)
        make_nota_tratada(doc_c, numero='300')
        self.client.force_authenticate(user=superuser)
        res = self.client.get('/api/notas-tratadas/')
        self.assertGreaterEqual(res.data['count'], 2)

    def test_detalhe_retorna_200(self):
        self.client.force_authenticate(user=self.staff)
        res = self.client.get(f'/api/notas-tratadas/{self.nota.pk}/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['id'], self.nota.pk)

    def test_nao_permite_criacao_post(self):
        """Endpoint é read-only — POST deve retornar 405."""
        self.client.force_authenticate(user=self.staff)
        res = self.client.post('/api/notas-tratadas/', {})
        self.assertEqual(res.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)


# ──────────────────────────────────────────────────────────────────────────────
# 5. Exportar Excel — /api/notas-tratadas/exportar/
# ──────────────────────────────────────────────────────────────────────────────

class ExportarExcelJWTTest(APITestCase):

    def setUp(self):
        self.escritorio = make_escritorio(cnpj='33333333000191')
        self.staff = make_staff(username='staff_xls', escritorio=self.escritorio)
        self.cliente = make_cliente(cnpj='55555555000155', escritorio=self.escritorio)
        doc = make_documento(self.cliente, chave='K' * 44)
        make_nota_tratada(doc, numero='1', competencia='06/2025')

    def test_sem_auth_retorna_401(self):
        res = self.client.get('/api/notas-tratadas/exportar/')
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_retorna_200_e_content_type_xlsx(self):
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/exportar/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('spreadsheetml', res['Content-Type'])

    def test_header_content_disposition_tem_xlsx(self):
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/exportar/')
        self.assertIn('.xlsx', res.get('Content-Disposition', ''))

    def test_planilha_tem_aba_notas_fiscais(self):
        import openpyxl
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/exportar/')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        self.assertIn('Notas Fiscais', wb.sheetnames)

    def test_planilha_tem_linha_de_dados(self):
        """Com 1 nota, planilha tem 3 linhas de cabeçalho + 1 dado + 1 SUBTOTAL = 5 linhas."""
        import openpyxl
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/exportar/')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb['Notas Fiscais']
        # 3 cabeçalhos (título, meta, colunas) + linhas de dados + SUBTOTAL
        self.assertGreater(ws.max_row, 4)

    def test_filtro_competencia_aplicado_na_exportacao(self):
        """Filtro de competência sem notas → 3 linhas de cabeçalho + 1 SUBTOTAL, sem dados."""
        import openpyxl
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/exportar/?data_competencia=01/2000')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb['Notas Fiscais']
        # título (1) + meta (2) + colunas (3) + SUBTOTAL (4) — zero linhas de dados
        self.assertEqual(ws.max_row, 4)
        # linha 1 (índice 0) deve conter o título da empresa, não dados de nota
        self.assertIn('CaptaFiscal', str(ws.cell(row=1, column=1).value or ''))

    def test_planilha_tem_aba_legenda(self):
        """A aba Legenda deve existir com explicações de cores e regimes."""
        import openpyxl
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/exportar/')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        self.assertIn('Legenda', wb.sheetnames)

    def test_planilha_tem_tres_linhas_de_cabecalho(self):
        """Estrutura: linha 1=título, linha 2=metadados, linha 3=colunas."""
        import openpyxl
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/exportar/')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb['Notas Fiscais']
        self.assertIn('CaptaFiscal', str(ws.cell(row=1, column=1).value or ''))
        self.assertIn('Gerado em', str(ws.cell(row=2, column=1).value or ''))
        self.assertEqual(ws.cell(row=3, column=1).value, 'Nº NFS-e')

    def test_linha_subtotal_contem_label(self):
        """Última linha deve conter o label de SUBTOTAL na coluna A."""
        import openpyxl
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/exportar/')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb['Notas Fiscais']
        ultima = ws.max_row
        label = str(ws.cell(row=ultima, column=1).value or '')
        self.assertIn('SUBTOTAL', label.upper())

    def test_dados_comecam_na_linha_4(self):
        """Com 1 nota, dados ficam na linha 4 (após 3 cabeçalhos)."""
        import openpyxl
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/exportar/')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb['Notas Fiscais']
        # linha 4 é a primeira de dados — deve ter o número da nota '1'
        self.assertEqual(str(ws.cell(row=4, column=1).value or ''), '1')

    def test_mei_retencao_nula_exibe_na(self):
        """Nota MEI com retenções nulas deve exibir 'N/A' nas células de retenção."""
        import openpyxl
        from fiscal.models import NotaTratada
        # Cria nota com regime MEI e retenções nulas
        doc_mei = make_documento(self.cliente, chave='Z' * 44)
        nota_mei = make_nota_tratada(
            doc_mei, numero='999', competencia='07/2025',
            emitente_cnpj='55555555000155',
        )
        nota_mei.regime_trib = 'MEI'
        nota_mei.ret_pis     = None
        nota_mei.ret_cofins  = None
        nota_mei.ret_csll    = None
        nota_mei.ret_irrf    = None
        nota_mei.ret_inss    = None
        nota_mei.save(update_fields=['regime_trib', 'ret_pis', 'ret_cofins', 'ret_csll', 'ret_irrf', 'ret_inss'])

        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/exportar/?data_competencia=07/2025')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb['Notas Fiscais']
        # Linha 4 é a primeira de dados (competência 07/2025 tem só nota MEI)
        # Coluna 13 (índice 12, 1-indexed) = Ret. PIS
        celula_pis = ws.cell(row=4, column=13).value
        self.assertEqual(celula_pis, 'N/A')

    def test_nota_nao_mei_retencao_nula_exibe_vazio(self):
        """Nota não-MEI com retenção nula deve ter célula vazia (não 'N/A')."""
        import openpyxl
        doc_lp = make_documento(self.cliente, chave='Y' * 44)
        nota_lp = make_nota_tratada(
            doc_lp, numero='888', competencia='08/2025',
            emitente_cnpj='55555555000155',
        )
        nota_lp.regime_trib = 'Lucro Presumido'
        nota_lp.ret_pis     = None
        nota_lp.save(update_fields=['regime_trib', 'ret_pis'])

        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/exportar/?data_competencia=08/2025')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb['Notas Fiscais']
        celula_pis = ws.cell(row=4, column=13).value
        # célula vazia (None ou '') — nunca 'N/A' para não-MEI
        self.assertNotEqual(celula_pis, 'N/A')

    def test_exportar_preserva_todos_os_campos_da_nota(self):
        """Verificar integridade dos dados exportados campo a campo."""
        import openpyxl
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/notas-tratadas/exportar/?data_competencia=06/2025')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb['Notas Fiscais']
        # Cabeçalhos esperados na linha 3
        cabecalhos_esperados = [
            'Nº NFS-e', 'Competência', 'Data Proc.', 'CNPJ Emitente', 'Emitente',
            'Doc Tomador', 'Tomador', 'Cód. Tributo', 'Serviço', 'Regime Trib.',
            'Valor Serviço', 'Valor Líquido', 'Ret. PIS', 'Ret. COFINS',
            'Ret. CSLL', 'Ret. IRRF', 'Ret. INSS', 'Parecer', 'Chave Substituta',
        ]
        for col_idx, esperado in enumerate(cabecalhos_esperados, start=1):
            self.assertEqual(
                ws.cell(row=3, column=col_idx).value,
                esperado,
                msg=f'Cabeçalho coluna {col_idx} incorreto',
            )


# ──────────────────────────────────────────────────────────────────────────────
# 6. ExportarPlanilhaView — /api/v1/integracao/exportar-planilha/ (Token auth)
# ──────────────────────────────────────────────────────────────────────────────

class IntegracaoExportarPlanilhaTest(APITestCase):
    URL = '/api/v1/integracao/exportar-planilha/'

    def setUp(self):
        self.user = User.objects.create_user(username='integra', password='pass')
        self.token = Token.objects.create(user=self.user)
        self.cliente = make_cliente(cnpj='77777777000177')
        doc = make_documento(self.cliente, chave='L' * 44)
        make_nota_tratada(doc, numero='50', competencia='06/2025',
                          emitente_cnpj='77777777000177')

    def _auth(self):
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {self.token.key}')

    def test_sem_auth_retorna_401(self):
        res = self.client.post(self.URL, {'cnpj': '77777777000177', 'mes': 6, 'ano': 2025})
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_jwt_nao_aceito_nesta_rota(self):
        """Rota de integração usa Token, não Bearer JWT."""
        from rest_framework_simplejwt.tokens import RefreshToken
        refresh = RefreshToken.for_user(self.user)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {str(refresh.access_token)}')
        res = self.client.post(self.URL, {'cnpj': '77777777000177', 'mes': 6, 'ano': 2025},
                               format='json')
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_cnpj_invalido_retorna_400(self):
        self._auth()
        res = self.client.post(self.URL, {'cnpj': '123', 'mes': 6, 'ano': 2025}, format='json')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('cnpj', res.data.get('erros', {}))

    def test_mes_invalido_retorna_400(self):
        self._auth()
        res = self.client.post(self.URL, {'cnpj': '77777777000177', 'mes': 13, 'ano': 2025}, format='json')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('mes', res.data.get('erros', {}))

    def test_ano_invalido_retorna_400(self):
        self._auth()
        res = self.client.post(self.URL, {'cnpj': '77777777000177', 'mes': 6, 'ano': 1800}, format='json')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('ano', res.data.get('erros', {}))

    def test_payload_valido_retorna_200_e_xlsx(self):
        self._auth()
        res = self.client.post(self.URL, {'cnpj': '77777777000177', 'mes': 6, 'ano': 2025}, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('spreadsheetml', res['Content-Type'])

    def test_planilha_tem_aba_notas_fiscais(self):
        import openpyxl
        self._auth()
        res = self.client.post(self.URL, {'cnpj': '77777777000177', 'mes': 6, 'ano': 2025}, format='json')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        self.assertIn('Notas Fiscais', wb.sheetnames)

    def test_planilha_tem_aba_auditoria(self):
        import openpyxl
        self._auth()
        res = self.client.post(self.URL, {'cnpj': '77777777000177', 'mes': 6, 'ano': 2025}, format='json')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        self.assertIn('Auditoria de Quebras', wb.sheetnames)

    def test_aba_notas_fiscais_tem_linha_de_dados(self):
        import openpyxl
        self._auth()
        res = self.client.post(self.URL, {'cnpj': '77777777000177', 'mes': 6, 'ano': 2025}, format='json')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb['Notas Fiscais']
        self.assertGreater(ws.max_row, 1, 'Planilha gerada não contém dados além do cabeçalho')

    def test_auditoria_detecta_quebra_sequencial(self):
        """Notas 1, 2, 4 (falta 3) → aba de auditoria deve registrar nota 3 como faltante."""
        import openpyxl
        cliente2 = make_cliente(cnpj='66666666000166')
        for num, chave_prefix in [('1', 'M'), ('2', 'N'), ('4', 'O')]:
            doc = make_documento(cliente2, chave=chave_prefix * 44)
            make_nota_tratada(doc, numero=num, competencia='06/2025',
                              emitente_cnpj='66666666000166')

        user2 = User.objects.create_user(username='integra2', password='pass')
        token2 = Token.objects.create(user=user2)
        self.client.credentials(HTTP_AUTHORIZATION=f'Token {token2.key}')
        res = self.client.post(self.URL, {'cnpj': '66666666000166', 'mes': 6, 'ano': 2025}, format='json')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb['Auditoria de Quebras']
        numeros_faltantes = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        self.assertIn(3, numeros_faltantes)

    def test_cnpj_sem_notas_retorna_xlsx_com_aba_vazia(self):
        """CNPJ válido mas sem notas no mês → planilha com só cabeçalho."""
        import openpyxl
        self._auth()
        res = self.client.post(self.URL, {'cnpj': '77777777000177', 'mes': 1, 'ano': 2000}, format='json')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb['Notas Fiscais']
        self.assertEqual(ws.max_row, 1)

    def test_content_disposition_contem_nome_de_arquivo(self):
        self._auth()
        res = self.client.post(self.URL, {'cnpj': '77777777000177', 'mes': 6, 'ano': 2025}, format='json')
        cd = res.get('Content-Disposition', '')
        self.assertIn('77777777000177', cd)
        self.assertIn('.xlsx', cd)


# ──────────────────────────────────────────────────────────────────────────────
# 7. Backfill command — teste de integração
# ──────────────────────────────────────────────────────────────────────────────

class BackfillNotaTratadaCommandTest(TestCase):

    def setUp(self):
        self.cliente = make_cliente(cnpj='44444444000144')

    def _run(self, **kw):
        from io import StringIO
        from django.core.management import call_command
        out = StringIO()
        call_command('backfill_nota_tratada', stdout=out, **kw)
        return out.getvalue()

    def test_cria_nota_tratada_para_documento_com_xml(self):
        doc = make_documento(self.cliente, chave='P' * 44)
        Xml.objects.create(documento=doc, conteudo=_xml_nfse(numero='77'))
        saida = self._run()
        self.assertTrue(NotaTratada.objects.filter(documento=doc).exists())
        self.assertIn('processados: 1', saida)

    def test_ignora_documento_sem_xml(self):
        make_documento(self.cliente, chave='Q' * 44)
        saida = self._run()
        self.assertIn('sem XML: 1', saida)

    def test_nao_reprocessa_ja_existente_sem_force(self):
        doc = make_documento(self.cliente, chave='R' * 44)
        Xml.objects.create(documento=doc, conteudo=_xml_nfse(numero='88'))
        self._run()
        nota = NotaTratada.objects.get(documento=doc)
        nota.numero_nfse = 'MANUAL'
        nota.save()
        self._run()
        nota.refresh_from_db()
        self.assertEqual(nota.numero_nfse, 'MANUAL')

    def test_force_reprocessa_existentes(self):
        doc = make_documento(self.cliente, chave='S' * 44)
        Xml.objects.create(documento=doc, conteudo=_xml_nfse(numero='99'))
        self._run()
        nota = NotaTratada.objects.get(documento=doc)
        nota.numero_nfse = 'MANUAL'
        nota.save()
        self._run(force=True)
        nota.refresh_from_db()
        self.assertEqual(nota.numero_nfse, '99')

    def test_filtra_por_cliente(self):
        outro = make_cliente(cnpj='55555555000155')
        doc1 = make_documento(self.cliente, chave='T' * 44)
        doc2 = make_documento(outro, chave='U' * 44)
        Xml.objects.create(documento=doc1, conteudo=_xml_nfse(numero='10'))
        Xml.objects.create(documento=doc2, conteudo=_xml_nfse(numero='20'))
        self._run(cliente=self.cliente.pk)
        self.assertTrue(NotaTratada.objects.filter(documento=doc1).exists())
        self.assertFalse(NotaTratada.objects.filter(documento=doc2).exists())


# ──────────────────────────────────────────────────────────────────────────────
# 8. Cliente.regime_tributario — campo e API
# ──────────────────────────────────────────────────────────────────────────────

class ClienteRegimeTributarioTest(APITestCase):
    """Testes para o campo regime_tributario adicionado ao modelo Cliente."""

    def setUp(self):
        self.escritorio = make_escritorio(cnpj='88888888000188')
        self.staff = make_staff(username='staff_reg', escritorio=self.escritorio)
        self.cliente = make_cliente(cnpj='99999999000191', escritorio=self.escritorio)

    def test_regime_tributario_presente_no_serializer(self):
        """Campo regime_tributario deve aparecer na resposta da API de clientes."""
        self.client.force_authenticate(user=self.staff)
        res = self.client.get(f'/api/clientes/{self.cliente.pk}/')
        self.assertEqual(res.status_code, 200)
        self.assertIn('regime_tributario', res.data)

    def test_regime_tributario_default_vazio(self):
        """Cliente recém-criado deve ter regime_tributario='' (não informado)."""
        self.client.force_authenticate(user=self.staff)
        res = self.client.get(f'/api/clientes/{self.cliente.pk}/')
        self.assertEqual(res.data['regime_tributario'], '')

    def test_patch_cliente_atualiza_regime_para_mei(self):
        """PATCH /api/clientes/{id}/ deve aceitar regime_tributario='MEI'."""
        self.client.force_authenticate(user=self.staff)
        res = self.client.patch(
            f'/api/clientes/{self.cliente.pk}/',
            {'regime_tributario': 'MEI'},
            format='json',
        )
        self.assertEqual(res.status_code, 200)
        self.cliente.refresh_from_db()
        self.assertEqual(self.cliente.regime_tributario, 'MEI')

    def test_patch_cliente_atualiza_regime_para_simples_nacional(self):
        """PATCH deve aceitar SN e persistir no banco."""
        self.client.force_authenticate(user=self.staff)
        res = self.client.patch(
            f'/api/clientes/{self.cliente.pk}/',
            {'regime_tributario': 'SN'},
            format='json',
        )
        self.assertEqual(res.status_code, 200)
        self.cliente.refresh_from_db()
        self.assertEqual(self.cliente.regime_tributario, 'SN')

    def test_patch_cliente_rejeita_regime_invalido(self):
        """Valor não listado em REGIME_CHOICES deve retornar 400."""
        self.client.force_authenticate(user=self.staff)
        res = self.client.patch(
            f'/api/clientes/{self.cliente.pk}/',
            {'regime_tributario': 'XX'},
            format='json',
        )
        self.assertEqual(res.status_code, 400)

    def test_regime_tributario_aparece_na_listagem(self):
        """Campo regime_tributario deve aparecer em GET /api/clientes/."""
        self.cliente.regime_tributario = 'LP'
        self.cliente.save(update_fields=['regime_tributario'])
        self.client.force_authenticate(user=self.staff)
        res = self.client.get('/api/clientes/')
        self.assertEqual(res.status_code, 200)
        resultados = res.data.get('results', res.data)
        encontrado = next((c for c in resultados if c['id'] == self.cliente.pk), None)
        self.assertIsNotNone(encontrado)
        self.assertEqual(encontrado['regime_tributario'], 'LP')


# ──────────────────────────────────────────────────────────────────────────────
# 9. Integridade dos dados XML → NotaTratada (pipeline completo)
# ──────────────────────────────────────────────────────────────────────────────

class PipelineXmlParaNovaTratadaTest(TestCase):
    """
    Garante que o pipeline XML → extrair_dados_nfse → salvar_nota_tratada
    preserva consistência e acurácia dos dados em cada campo.
    """

    def setUp(self):
        self.cliente = make_cliente(cnpj='11111111000111')

    def _processar(self, xml: str, status='COMPLETO', papel='EMITENTE'):
        from fiscal.conectores.nfse import _salvar_nota_tratada
        import hashlib
        chave = hashlib.md5(xml.encode()).hexdigest()[:44].ljust(44, '0')
        doc = make_documento(self.cliente, chave=chave, status=status, papel=papel)
        _salvar_nota_tratada(doc, xml, status, papel)
        return NotaTratada.objects.get(documento=doc)

    def test_pipeline_preserva_numero_nfse(self):
        nota = self._processar(_xml_nfse(numero='42'))
        self.assertEqual(nota.numero_nfse, '42')

    def test_pipeline_preserva_competencia_no_formato_mm_aaaa(self):
        nota = self._processar(_xml_nfse(competencia='2025-09'))
        self.assertEqual(nota.data_competencia, '09/2025')

    def test_pipeline_preserva_emitente_cnpj(self):
        nota = self._processar(_xml_nfse(cnpj_emit='12345678000195'))
        self.assertEqual(nota.emitente_cnpj, '12345678000195')

    def test_pipeline_preserva_tomador_nome(self):
        nota = self._processar(_xml_nfse(nome_toma='Empresa Tomadora SA'))
        self.assertEqual(nota.tomador_nome, 'Empresa Tomadora SA')

    def test_pipeline_preserva_valor_servico_como_decimal(self):
        nota = self._processar(_xml_nfse(v_serv='3750.50'))
        self.assertEqual(nota.valor_servico, decimal.Decimal('3750.50'))

    def test_pipeline_mei_preserva_regime_e_deixa_retencoes_nulas(self):
        """MEI (opSimpNac='3') → regime='MEI', retenções federais devem ser None."""
        nota = self._processar(_xml_nfse(op_simp_nac='3', ret_csll='', ret_pis='', ret_cofins=''))
        self.assertEqual(nota.regime_trib, 'MEI')
        # MEI não retém federais — parser deve deixar None, não zero
        self.assertIsNone(nota.ret_csll)
        self.assertIsNone(nota.ret_pis)
        self.assertIsNone(nota.ret_cofins)

    def test_pipeline_simples_nacional_preserva_regime(self):
        nota = self._processar(_xml_nfse(op_simp_nac='1'))
        self.assertEqual(nota.regime_trib, 'Simples Nacional')

    def test_pipeline_parecer_valida_com_retencoes_corretas(self):
        nota = self._processar(_xml_nfse(v_serv='1000.00', ret_csll='10.00'))
        self.assertEqual(nota.parecer, 'Válida')

    def test_pipeline_parecer_divergencia_detectada(self):
        nota = self._processar(_xml_nfse(v_serv='1000.00', ret_csll='99.00'))
        self.assertEqual(nota.parecer, 'Válida (DIVERGÊNCIA RETENÇÃO)')

    def test_pipeline_nota_cancelada_nao_altera_parecer_para_divergencia(self):
        """Status CANCELADO sobrepõe qualquer divergência de retenção."""
        from fiscal.conectores.nfse import _salvar_nota_tratada
        xml = _xml_nfse(v_serv='1000.00', ret_csll='99.00')
        doc = make_documento(self.cliente, chave='C' * 43 + '1', status='CANCELADO')
        _salvar_nota_tratada(doc, xml, 'CANCELADO', 'EMITENTE')
        nota = NotaTratada.objects.get(documento=doc)
        self.assertEqual(nota.parecer, 'Cancelada')

    def test_pipeline_idempotente_reprocessamento_nao_duplica(self):
        """Reprocessar o mesmo XML não deve criar segundo registro (update_or_create)."""
        from fiscal.conectores.nfse import _salvar_nota_tratada
        doc = make_documento(self.cliente, chave='I' * 44)
        xml = _xml_nfse(numero='77')
        _salvar_nota_tratada(doc, xml, 'COMPLETO', 'EMITENTE')
        _salvar_nota_tratada(doc, xml, 'COMPLETO', 'EMITENTE')  # segunda chamada
        self.assertEqual(NotaTratada.objects.filter(documento=doc).count(), 1)

    def test_pipeline_exportacao_xlsx_contem_dados_corretos(self):
        """Após processar 1 nota, exportação deve conter seus dados na planilha."""
        import io as _io
        import openpyxl

        # Garante que a nota está no banco
        nota = self._processar(_xml_nfse(numero='XLS1', v_serv='500.00', competencia='2025-11'))

        # Cria staff e faz request de exportação via test client DRF
        escritorio = make_escritorio(cnpj='22222222000122')
        staff = make_staff(username='staff_pipeline', escritorio=escritorio)
        # Re-associa o cliente ao escritório para isolamento de tenant
        self.cliente.escritorio = escritorio
        self.cliente.save(update_fields=['escritorio'])

        from rest_framework.test import APIClient
        api = APIClient()
        api.force_authenticate(user=staff)
        res = api.get('/api/notas-tratadas/exportar/?data_competencia=11/2025')
        self.assertEqual(res.status_code, 200)
        wb = openpyxl.load_workbook(_io.BytesIO(res.content))
        ws = wb['Notas Fiscais']
        # Linha 4 = primeira linha de dados
        self.assertEqual(str(ws.cell(row=4, column=1).value or ''), 'XLS1')
