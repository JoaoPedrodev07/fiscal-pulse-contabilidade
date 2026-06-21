"""
Endpoint externo de integração para clientes que consomem a API via Token.

POST /api/v1/integracao/exportar-planilha/
    Body: { "cnpj": "12345678000199", "mes": 6, "ano": 2025 }
    Auth: Authorization: Token <api-key>
    Response: arquivo .xlsx com duas abas:
        "Notas Fiscais"       — dados tratados + parecer de todas as NFS-e do CNPJ/competência
        "Auditoria de Quebras"— gaps na sequência numérica das notas (possíveis fraudes/falhas)
"""
from __future__ import annotations

import io
import re

from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import TokenAuthentication
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status

from .models import NotaTratada, Documento


class ExportarPlanilhaView(APIView):
    """
    Exporta planilha Excel de NFS-e tratadas para integração externa.
    Autenticação via Token DRF (não JWT — apta para sistemas de terceiros).
    """
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        cnpj = request.data.get('cnpj', '').strip()
        mes  = request.data.get('mes')
        ano  = request.data.get('ano')

        erros = {}
        if not re.fullmatch(r'\d{14}', cnpj):
            erros['cnpj'] = 'CNPJ deve conter 14 dígitos numéricos, sem pontuação.'
        try:
            mes_int = int(mes)
            if not 1 <= mes_int <= 12:
                raise ValueError
        except (TypeError, ValueError):
            erros['mes'] = 'Mês inválido. Informe um inteiro de 1 a 12.'
            mes_int = 0
        try:
            ano_int = int(ano)
            if not 2000 <= ano_int <= 2100:
                raise ValueError
        except (TypeError, ValueError):
            erros['ano'] = 'Ano inválido. Informe um inteiro entre 2000 e 2100.'
            ano_int = 0

        if erros:
            return Response({'erros': erros}, status=status.HTTP_400_BAD_REQUEST)

        competencia_mm_aaaa = f'{mes_int:02d}/{ano_int}'

        notas_qs = (
            NotaTratada.objects
            .filter(emitente_cnpj=cnpj, data_competencia=competencia_mm_aaaa)
            .select_related('documento')
            .order_by('numero_nfse')
        )

        buf = io.BytesIO()
        _gerar_xlsx(buf, notas_qs, cnpj, competencia_mm_aaaa)
        buf.seek(0)

        filename = f'relatorio_{cnpj}_{ano_int}-{mes_int:02d}.xlsx'
        from django.http import HttpResponse
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


def _gerar_xlsx(buf: io.BytesIO, notas_qs, cnpj: str, competencia: str) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # ── Aba 1: Notas Fiscais ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Notas Fiscais'

    cabecalhos = [
        'Nº NFS-e', 'Competência', 'Data Proc.', 'CNPJ Emitente', 'Emitente',
        'Doc Tomador', 'Tomador', 'Cód. Tributo', 'Serviço', 'Regime Trib.',
        'Valor Serviço', 'Valor Líquido',
        'Ret. PIS', 'Ret. COFINS', 'Ret. CSLL', 'Ret. IRRF', 'Ret. INSS',
        'Parecer', 'Chave Substituta',
    ]

    header_fill = PatternFill('solid', fgColor='2563EB')
    header_font = Font(bold=True, color='FFFFFF')
    for col, titulo in enumerate(cabecalhos, start=1):
        cell = ws1.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    PARECER_CORES = {
        'Válida':                        'D1FAE5',
        'Válida (DIVERGÊNCIA RETENÇÃO)': 'FEF3C7',
        'Cancelada':                     'FEE2E2',
        'Substituída':                   'E0E7FF',
    }

    numeros_vistos: list[int] = []

    for row_idx, nota in enumerate(notas_qs, start=2):
        try:
            numeros_vistos.append(int(nota.numero_nfse))
        except (ValueError, TypeError):
            pass

        linha = [
            nota.numero_nfse,
            nota.data_competencia,
            nota.data_processamento,
            nota.emitente_cnpj,
            nota.emitente_nome,
            nota.tomador_doc,
            nota.tomador_nome,
            nota.codigo_tributo,
            nota.descricao_servico,
            nota.regime_trib,
            float(nota.valor_servico) if nota.valor_servico is not None else None,
            float(nota.valor_liquido) if nota.valor_liquido is not None else None,
            float(nota.ret_pis)    if nota.ret_pis    is not None else None,
            float(nota.ret_cofins) if nota.ret_cofins is not None else None,
            float(nota.ret_csll)   if nota.ret_csll   is not None else None,
            float(nota.ret_irrf)   if nota.ret_irrf   is not None else None,
            float(nota.ret_inss)   if nota.ret_inss   is not None else None,
            nota.parecer,
            nota.chave_substituta,
        ]
        for col_idx, valor in enumerate(linha, start=1):
            ws1.cell(row=row_idx, column=col_idx, value=valor)

        cor = PARECER_CORES.get(nota.parecer)
        if cor:
            fill = PatternFill('solid', fgColor=cor)
            for col_idx in range(1, len(cabecalhos) + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = fill

    for col in ws1.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # ── Aba 2: Auditoria de Quebras ───────────────────────────────────────────
    ws2 = wb.create_sheet(title='Auditoria de Quebras')

    cab2 = ['Nº Esperado', 'Nº Anterior', 'Nº Seguinte', 'Qtd. Faltando', 'Observação']
    header_fill2 = PatternFill('solid', fgColor='DC2626')
    for col, titulo in enumerate(cab2, start=1):
        cell = ws2.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill2
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    numeros_vistos.sort()
    gaps_row = 2
    if len(numeros_vistos) >= 2:
        for i in range(len(numeros_vistos) - 1):
            anterior = numeros_vistos[i]
            proximo  = numeros_vistos[i + 1]
            if proximo - anterior > 1:
                for faltando in range(anterior + 1, proximo):
                    ws2.cell(row=gaps_row, column=1, value=faltando)
                    ws2.cell(row=gaps_row, column=2, value=anterior)
                    ws2.cell(row=gaps_row, column=3, value=proximo)
                    ws2.cell(row=gaps_row, column=4, value=proximo - anterior - 1)
                    ws2.cell(row=gaps_row, column=5, value='NSU ausente — verificar cancelamento ou captura pendente')
                    gaps_row += 1

    if gaps_row == 2:
        ws2.cell(row=2, column=1, value='Nenhuma quebra detectada na sequência numérica.')

    for col in ws2.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=12)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(buf)
